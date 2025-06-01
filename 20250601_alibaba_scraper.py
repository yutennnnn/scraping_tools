import re
import os
import time
import math
import threading
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

APP_BG = "#181D23"
FRAME_BG = "#222831"
ENTRY_BG = "#323943"
ENTRY_FG = "#F5F6FA"
LABEL_FG = "#C8CCD6"
BTN_BG = "#4f8cff"
BTN_FG = "#222831"
BTN_HOVER_BG = "#6ea8fe"
BTN_ACTIVE_BG = "#2366d1"
GUIDE_FG = "#757575"
FONT_MAIN = ("Segoe UI", 13)
FONT_TITLE = ("Segoe UI", 19, "bold")

def normalize_option_name(text):
    if not text:
        return ""
    text = re.sub(r"[\s\u3000（）\[\]【】\(\)『』]", "", text)
    text = text.replace('&quot;', '').replace('"', '').replace("'", "")
    return text.strip()

def extract_price_1688(url, driver, option_value=None):
    try:
        driver.get(url)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "body"))
        )
        time.sleep(2.5)

        # SKUパターン1：従来のdiv.prop-name[title]
        prop_divs = driver.find_elements(By.CSS_SELECTOR, "div.prop-name[title]")
        sku_set = set()
        for div in prop_divs:
            title = div.get_attribute("title")
            norm_title = normalize_option_name(title)
            if norm_title:
                sku_set.add(norm_title)

        # SKUパターン2：テーブル型（.sku-prop-module-name+div .sku-item-wrapperなど）も取得
        sku_table_rows = driver.find_elements(By.CSS_SELECTOR, ".sku-prop-module-name + div .sku-item-wrapper")
        for tr in sku_table_rows:
            txt = tr.text.strip()
            if txt:
                sku_set.add(txt)

        # C列空欄かつSKU2種以上 → 絶対に金額抽出せず
        if not (option_value and str(option_value).strip()):
            if len(sku_set) >= 2:
                return "選択肢未入力価格無し"
            else:
                try:
                    price_block = driver.find_element(By.CSS_SELECTOR, "div.discountPrice-price")
                    txt = price_block.text.strip()
                    m = re.search(r"([0-9]+(?:\.[0-9]+)?)元", txt)
                    if m:
                        return m.group(1)
                except Exception as e:
                    print(f"[WARN] 価格エリア取得失敗: {e}")
                html = driver.page_source
                m3 = re.search(r"([0-9]+(?:\.[0-9]+)?)元", html)
                if m3:
                    return m3.group(1)
                return "URLエラー"

        # C列指定時：従来のSKUクリック
        norm_opt = normalize_option_name(option_value)
        matched_div = None
        for div in prop_divs:
            title = div.get_attribute("title")
            if title and (norm_opt == normalize_option_name(title) or norm_opt in normalize_option_name(title)):
                matched_div = div
                break
        if matched_div:
            try:
                driver.execute_script("arguments[0].scrollIntoView(true);", matched_div)
                time.sleep(0.2)
                matched_div.click()
                time.sleep(1.0)
                WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "div.discountPrice-price"))
                )
                price_block = driver.find_element(By.CSS_SELECTOR, "div.discountPrice-price")
                txt = price_block.text.strip()
                m = re.search(r"([0-9]+(?:\.[0-9]+)?)元", txt)
                if m:
                    return m.group(1)
            except Exception as e:
                print(f"[WARN] 選択肢クリックor価格抽出失敗: {e}")

        return "URLエラー"
    except Exception as e:
        print(f"[ERROR] extract_price_1688 {url}: {e}")
        return "URLエラー"

def process_excel(excel_path, gui_status_callback=None, progress_var=None, total_var=None):
    wb = load_workbook(excel_path)
    ws = wb.active

    header_row = 1
    url_col = 2
    option_col = 3
    price_col = 11
    min_row = header_row + 1
    last_row = ws.max_row

    wait_per_row = 4
    total_rows = last_row - min_row + 1
    t0 = time.time()
    processed = 0

    options = Options()
    options.add_argument('--headless=new')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(10)

    for row in range(min_row, last_row + 1):
        url_val = ws.cell(row=row, column=url_col).value
        option_val = ws.cell(row=row, column=option_col).value

        if url_val is None or str(url_val).strip() == "":
            ws.cell(row=row, column=price_col).value = None
            continue

        url = str(url_val).strip()
        option = str(option_val).strip() if option_val else None

        processed += 1
        remaining = total_rows - processed
        eta = max(0, math.ceil(remaining * wait_per_row))
        status_msg = f"進捗: {processed}/{total_rows}件 | 残り想定: 約{eta}秒\n{url} の価格取得中…"

        if gui_status_callback:
            gui_status_callback(status_msg)

        if progress_var and total_var:
            progress_var.set(processed)
            total_var.set(total_rows)

        price = extract_price_1688(url, driver, option_value=option)
        ws.cell(row=row, column=price_col).value = price

        if gui_status_callback:
            gui_status_callback(status_msg + f"\n抽出結果: {price}")

        time.sleep(1.5)

    driver.quit()
    wb.save(excel_path)
    wb.close()
    return processed

def main():
    def select_excel():
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        excel_path_var.set(path)

    def set_status(msg):
        status_var.set(msg)
        status_label.update_idletasks()

    def run_scraping():
        excel_path = excel_path_var.get()
        if not os.path.exists(excel_path):
            messagebox.showerror("エラー", "Excelファイルを選択してください")
            return
        btn_exec.config(state="disabled")
        progress_bar["maximum"] = 1  # 仮
        progress_var.set(0)
        total_var.set(1)
        def scraping_thread():
            set_status("処理開始...")
            try:
                wb = load_workbook(excel_path)
                ws = wb.active
                header_row = 1
                min_row = header_row + 1
                last_row = ws.max_row
                total = last_row - min_row + 1
                progress_bar["maximum"] = total
                total_var.set(total)
                process_excel(
                    excel_path, gui_status_callback=set_status,
                    progress_var=progress_var, total_var=total_var
                )
                set_status(f"すべて完了！（処理件数: {total_var.get()}）")
                messagebox.showinfo("完了", f"すべてのURLの金額抽出が完了しました。")
            except Exception as e:
                set_status(f"エラー: {e}")
                messagebox.showerror("エラー", str(e))
            btn_exec.config(state="normal")
        threading.Thread(target=scraping_thread, daemon=True).start()

    root = tk.Tk()
    root.title("1688 Price Scraper")
    root.configure(bg=APP_BG)
    root.geometry("650x370")
    root.resizable(False, False)

    title = tk.Label(
        root, text="1688 Price Scraper", font=FONT_TITLE, fg="#85b5ff", bg=APP_BG, pady=15
    )
    title.pack()

    frame = tk.Frame(root, bg=FRAME_BG, bd=0, relief=tk.RIDGE)
    frame.pack(pady=12, padx=24, fill="both", expand=True)

    tk.Label(frame, text="Excelファイル（.xlsx）", font=FONT_MAIN, fg=LABEL_FG, bg=FRAME_BG, anchor="w").pack(
        pady=(20, 2), padx=18, anchor="w"
    )
    excel_path_var = tk.StringVar()
    entry_excel = tk.Entry(
        frame, textvariable=excel_path_var, font=FONT_MAIN, bg=ENTRY_BG, fg=ENTRY_FG,
        bd=1, relief=tk.FLAT, insertbackground=ENTRY_FG
    )
    entry_excel.pack(padx=18, fill="x")
    btn_file = tk.Button(
        frame, text="ファイル選択", font=FONT_MAIN, command=select_excel,
        bg=BTN_BG, fg=BTN_FG, activebackground=BTN_ACTIVE_BG, activeforeground=BTN_FG, bd=0, relief=tk.FLAT, cursor="hand2"
    )
    btn_file.pack(padx=18, pady=5, anchor="e")
    btn_file.bind("<Enter>", lambda e: btn_file.config(bg=BTN_HOVER_BG))
    btn_file.bind("<Leave>", lambda e: btn_file.config(bg=BTN_BG))
    btn_file.bind("<ButtonPress-1>", lambda e: btn_file.config(bg=BTN_ACTIVE_BG))
    btn_file.bind("<ButtonRelease-1>", lambda e: btn_file.config(bg=BTN_HOVER_BG))

    global btn_exec
    btn_exec = tk.Button(
        frame, text="スクレイピング", font=FONT_MAIN,
        command=run_scraping, bg=BTN_BG, fg=BTN_FG,
        activebackground=BTN_ACTIVE_BG, activeforeground=BTN_FG,
        height=2, bd=0, relief=tk.FLAT, cursor="hand2"
    )
    btn_exec.pack(pady=18, padx=18, fill="x")
    btn_exec.bind("<Enter>", lambda e: btn_exec.config(bg=BTN_HOVER_BG))
    btn_exec.bind("<Leave>", lambda e: btn_exec.config(bg=BTN_BG))
    btn_exec.bind("<ButtonPress-1>", lambda e: btn_exec.config(bg=BTN_ACTIVE_BG))
    btn_exec.bind("<ButtonRelease-1>", lambda e: btn_exec.config(bg=BTN_HOVER_BG))

    # プログレスバーと進捗テキスト
    progress_var = tk.IntVar(value=0)
    total_var = tk.IntVar(value=1)
    progress_frame = tk.Frame(root, bg=APP_BG)
    progress_frame.pack(pady=(0, 8))
    progress_bar = ttk.Progressbar(
        progress_frame, length=500, variable=progress_var, maximum=total_var.get(), mode="determinate"
    )
    progress_bar.pack(side="left", padx=(20, 0))
    progress_label = tk.Label(
        progress_frame, textvariable=progress_var, font=("Segoe UI", 11),
        fg="#77aaff", bg=APP_BG, width=4, anchor="w"
    )
    progress_label.pack(side="left", padx=(8, 0))
    tk.Label(progress_frame, text="/", font=("Segoe UI", 11), fg="#77aaff", bg=APP_BG).pack(side="left")
    total_label = tk.Label(
        progress_frame, textvariable=total_var, font=("Segoe UI", 11),
        fg="#77aaff", bg=APP_BG, width=4, anchor="w"
    )
    total_label.pack(side="left")

    status_var = tk.StringVar()
    status_var.set("Excelファイルを選択し、ボタンを押してください。")
    status_label = tk.Label(frame, textvariable=status_var, font=("Segoe UI", 11), fg="#77aaff", bg=FRAME_BG, anchor="w", wraplength=570, justify="left")
    status_label.pack(pady=(7, 10), padx=18, anchor="w")

    guide = tk.Label(
        root,
        text="B列URL・C列選択肢の商品価格をK列に自動出力＆進捗バー表示します。",
        font=("Segoe UI", 10), fg=GUIDE_FG, bg=APP_BG
    )
    guide.pack(side="bottom", pady=7)

    root.mainloop()

if __name__ == "__main__":
    main()