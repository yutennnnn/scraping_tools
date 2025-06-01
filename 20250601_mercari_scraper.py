import re
import os
import time
import math
import threading
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# Optional private helpers or data
try:
    from private_code import *  # noqa: F401,F403
except Exception:
    pass

# UI設定
APP_BG = "#181D23"
FRAME_BG = "#222831"
ENTRY_BG = "#323943"
ENTRY_FG = "#F5F6FA"
LABEL_FG = "#C8CCD6"
BTN_BG = "#4f8cff"
BTN_FG = "#222831"
BTN_HOVER_BG = "#6ea8fe"
BTN_ACTIVE_BG = "#2366d1"
GUIDE_FG = "#757575"
FONT_MAIN = ("Segoe UI", 13)
FONT_TITLE = ("Segoe UI", 19, "bold")

def extract_price(url, driver):
    try:
        driver.get(url)
        time.sleep(2)

        item_price_spans = driver.find_elements(By.CSS_SELECTOR, 'div[data-testid="price"] > span')
        if len(item_price_spans) >= 2:
            price = re.sub(r'[^\d]', '', item_price_spans[1].text)
            if price:
                return price

        shop_price_spans = driver.find_elements(By.CSS_SELECTOR, 'div[data-testid="product-price"] > span')
        if len(shop_price_spans) >= 2:
            price = re.sub(r'[^\d]', '', shop_price_spans[1].text)
            if price:
                return price

        html = driver.page_source
        m = re.search(r'itemprop="price"[^>]*content="(\d+)"', html)
        if m:
            return m.group(1)
        m2 = re.search(r'<span[^>]*itemprop="price"[^>]*>([\d,]+)</span>', html)
        if m2:
            return m2.group(1).replace(',', '')

        return "URLエラー"
    except Exception as e:
        print(f"[ERROR] extract_price {url}: {e}")
        return "URLエラー"

def process_excel(excel_path, gui_status_callback=None, progress_var=None, total_var=None, progress_bar=None):
    wb = load_workbook(excel_path)
    ws = wb.active

    header_row = 1     # <<<<<< 必ず1に！（1行目=ヘッダー、2行目=最初のデータ行）
    url_col = 4        # D列
    price_col = 12     # L列
    min_row = header_row + 1
    last_row = ws.max_row

    wait_per_row = 3
    total_rows = last_row - min_row + 1 + 1  # <<<<<< +1で2行目も進捗バー含む
    processed = 0

    if progress_bar:
        progress_bar["maximum"] = total_rows

    options = Options()
    options.add_argument('--headless=new')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(10)

    t0 = time.time()

    for row in range(min_row, last_row + 1):
        cell_val = ws.cell(row=row, column=url_col).value
        print(f"[DEBUG] row={row}, D列値={repr(cell_val)}")

        if cell_val is None or str(cell_val).strip() == "":
            ws.cell(row=row, column=price_col).value = None
            if progress_var:
                progress_var.set(processed)
                if progress_bar: progress_bar.update()
            continue

        cell_str = str(cell_val).strip()
        urls = re.findall(r'https://jp\.mercari\.com/(?:item|shops/product)/[a-zA-Z0-9]+', cell_str)
        price = None

        processed += 1
        remaining = total_rows - processed
        eta = max(0, math.ceil(remaining * wait_per_row))
        status_msg = f"進捗: {processed}/{total_rows}件 | 残り想定: 約{eta}秒"

        if urls:
            for url in urls:
                print(f"[INFO] ({row}) Try: {url}")
                if gui_status_callback:
                    gui_status_callback(f"{status_msg}\n行{row}: {url}の価格取得中…")
                price = extract_price(url, driver)
                if price and price not in ["URLエラー"]:
                    ws.cell(row=row, column=price_col).value = int(price)
                    print(f"[SUCCESS] 行{row}: {url} → {price}円")
                    if gui_status_callback:
                        gui_status_callback(f"{status_msg}\n行{row}: {url} → {price}円")
                    break
            if not price or price == "URLエラー":
                ws.cell(row=row, column=price_col).value = price
                print(f"[FAIL] 行{row}: {url} → {price}")
                if gui_status_callback:
                    gui_status_callback(f"{status_msg}\n行{row}: {url} → {price}")
        else:
            ws.cell(row=row, column=price_col).value = "選択肢未入力価格無し"
            print(f"[SKIP] 行{row}: 商品URLがD列に見つからず")
            if gui_status_callback:
                gui_status_callback(f"{status_msg}\n行{row}: 商品URLがD列に見つからず")

        if progress_var:
            progress_var.set(processed)
            if progress_bar: progress_bar.update()

        time.sleep(1.5)

    driver.quit()
    wb.save(excel_path)
    wb.close()
    print(f"--- 完了：処理行数={processed} ---")
    return processed

def main():
    def select_excel():
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        excel_path_var.set(path)

    def set_status(msg):
        status_var.set(msg)
        status_label.update_idletasks()

    def run_scraping():
        excel_path = excel_path_var.get()
        if not os.path.exists(excel_path):
            messagebox.showerror("エラー", "Excelファイルを選択してください")
            return
        btn_exec.config(state="disabled")
        progress_bar["maximum"] = 1  # 仮
        progress_var.set(0)
        total_var.set(1)
        def scraping_thread():
            set_status("処理開始...")
            try:
                wb = load_workbook(excel_path)
                ws = wb.active
                header_row = 1  # <<<<<< ここも1に
                min_row = header_row + 1
                last_row = ws.max_row
                total = last_row - min_row + 1 + 1  # +1で2行目も進捗バー含む
                progress_bar["maximum"] = total
                total_var.set(total)
                process_excel(
                    excel_path, gui_status_callback=set_status,
                    progress_var=progress_var, total_var=total_var, progress_bar=progress_bar
                )
                set_status(f"すべて完了！（処理件数: {total_var.get()}）")
                messagebox.showinfo("完了", f"すべてのURLの金額抽出が完了しました。")
            except Exception as e:
                set_status(f"エラー: {e}")
                messagebox.showerror("エラー", str(e))
            btn_exec.config(state="normal")
        threading.Thread(target=scraping_thread, daemon=True).start()

    root = tk.Tk()
    root.title("Mercari Price Extractor")
    root.configure(bg=APP_BG)
    root.geometry("650x370")
    root.resizable(False, False)

    title = tk.Label(
        root, text="Mercari Price Extractor", font=FONT_TITLE, fg="#85b5ff", bg=APP_BG, pady=15
    )
    title.pack()

    frame = tk.Frame(root, bg=FRAME_BG, bd=0, relief=tk.RIDGE)
    frame.pack(pady=12, padx=24, fill="both", expand=True)

    tk.Label(frame, text="Excelファイル（.xlsx）", font=FONT_MAIN, fg=LABEL_FG, bg=FRAME_BG, anchor="w").pack(
        pady=(20, 2), padx=18, anchor="w"
    )
    excel_path_var = tk.StringVar()
    entry_excel = tk.Entry(
        frame, textvariable=excel_path_var, font=FONT_MAIN, bg=ENTRY_BG, fg=ENTRY_FG,
        bd=1, relief=tk.FLAT, insertbackground=ENTRY_FG
    )
    entry_excel.pack(padx=18, fill="x")
    btn_file = tk.Button(
        frame, text="ファイル選択", font=FONT_MAIN, command=select_excel,
        bg=BTN_BG, fg=BTN_FG, activebackground=BTN_ACTIVE_BG, activeforeground=BTN_FG, bd=0, relief=tk.FLAT, cursor="hand2"
    )
    btn_file.pack(padx=18, pady=5, anchor="e")
    btn_file.bind("<Enter>", lambda e: btn_file.config(bg=BTN_HOVER_BG))
    btn_file.bind("<Leave>", lambda e: btn_file.config(bg=BTN_BG))
    btn_file.bind("<ButtonPress-1>", lambda e: btn_file.config(bg=BTN_ACTIVE_BG))
    btn_file.bind("<ButtonRelease-1>", lambda e: btn_file.config(bg=BTN_HOVER_BG))

    global btn_exec
    btn_exec = tk.Button(
        frame, text="スクレイピング", font=FONT_MAIN,
        command=run_scraping, bg=BTN_BG, fg=BTN_FG,
        activebackground=BTN_ACTIVE_BG, activeforeground=BTN_FG,
        height=2, bd=0, relief=tk.FLAT, cursor="hand2"
    )
    btn_exec.pack(pady=18, padx=18, fill="x")
    btn_exec.bind("<Enter>", lambda e: btn_exec.config(bg=BTN_HOVER_BG))
    btn_exec.bind("<Leave>", lambda e: btn_exec.config(bg=BTN_BG))
    btn_exec.bind("<ButtonPress-1>", lambda e: btn_exec.config(bg=BTN_ACTIVE_BG))
    btn_exec.bind("<ButtonRelease-1>", lambda e: btn_exec.config(bg=BTN_HOVER_BG))

    # プログレスバーと進捗テキスト
    progress_var = tk.IntVar(value=0)
    total_var = tk.IntVar(value=1)
    progress_frame = tk.Frame(root, bg=APP_BG)
    progress_frame.pack(pady=(0, 8))
    progress_bar = ttk.Progressbar(
        progress_frame, length=500, variable=progress_var, maximum=total_var.get(), mode="determinate"
    )
    progress_bar.pack(side="left", padx=(20, 0))
    progress_label = tk.Label(
        progress_frame, textvariable=progress_var, font=("Segoe UI", 11),
        fg="#77aaff", bg=APP_BG, width=4, anchor="w"
    )
    progress_label.pack(side="left", padx=(8, 0))
    tk.Label(progress_frame, text="/", font=("Segoe UI", 11), fg="#77aaff", bg=APP_BG).pack(side="left")
    total_label = tk.Label(
        progress_frame, textvariable=total_var, font=("Segoe UI", 11),
        fg="#77aaff", bg=APP_BG, width=4, anchor="w"
    )
    total_label.pack(side="left")

    status_var = tk.StringVar()
    status_var.set("Excelファイルを選択し、ボタンを押してください。")
    status_label = tk.Label(frame, textvariable=status_var, font=("Segoe UI", 11), fg="#77aaff", bg=FRAME_BG, anchor="w", wraplength=570, justify="left")
    status_label.pack(pady=(7, 10), padx=18, anchor="w")

    guide = tk.Label(
        root,
        text="D列URL巡回・L列へ金額出力",
        font=("Segoe UI", 10), fg=GUIDE_FG, bg=APP_BG
    )
    guide.pack(side="bottom", pady=7)

    root.mainloop()

if __name__ == "__main__":
    main()